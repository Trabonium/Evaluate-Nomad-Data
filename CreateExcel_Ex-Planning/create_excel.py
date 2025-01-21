import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime


class ExperimentExcelBuilder:
    def __init__(self, process_sequence, process_config, first_row_color="FFFF00", second_row_color="F77C00"):
        self.process_sequence = process_sequence
        self.process_config = process_config
        self.first_row_fill = PatternFill(
            start_color=first_row_color, end_color=first_row_color, fill_type="solid")
        self.second_row_fill = PatternFill(
            start_color=second_row_color, end_color=second_row_color, fill_type="solid")
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def build_excel(self):
        start_col = 1
        incremental_number = 1

        for process_data in self.process_sequence:
            process_name = process_data['process']
            custom_config = process_data.get('config', {})

            # Generate the dynamic steps for this process based on the configuration
            steps = self.generate_steps_for_process(process_name, custom_config)

            step_count = len(steps)
            end_col = start_col + step_count - 1

            # Determine the label for the process
            process_label = f"{incremental_number}: {process_name}" if process_name != "Experiment Info" else process_name
            if process_name != "Experiment Info":
                incremental_number += 1

            # Merge the cells in the first row for the process
            self.worksheet.merge_cells(
                start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            # Set the process name in the merged cell, align it to the center, and apply color
            cell = self.worksheet.cell(row=1, column=start_col)
            cell.value = process_label
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = self.first_row_fill

            # Fill in the steps in the second row and apply color
            for i, step in enumerate(steps):
                cell = self.worksheet.cell(row=2, column=start_col + i)
                cell.value = step
                cell.fill = self.second_row_fill

            # Move to the next set of columns
            start_col = end_col + 1

        # Apply the custom formula for the "Nomad ID" column
        self.apply_nomad_id_formula()

        # Adjust the width of the columns based on their contents
        self.adjust_column_widths()

    def generate_steps_for_process(self, process_name, custom_config):
        # Use default process config if custom_config is not provided
        config = self.process_config.get(process_name, {}).copy()
        config.update(custom_config)  # Override with custom config if provided

        if process_name == "Cleaning O2-Plasma" or process_name == "Cleaning UV-Ozone":
            steps = []
            for i in range(1, config.get("solvents", 0) + 1):
                steps.extend([f"Solvent {i}", f"Time {i} [s]",
                             f"Temperature {i} [°C]"])
            if process_name == "Cleaning O2-Plasma":
                steps.extend(
                    ["Gas-Plasma Gas", "Gas-Plasma Time [s]", "Gas-Plasma Power [W]"])
            if process_name == "Cleaning UV-Ozone":
                steps.extend(["UV-Ozone Time [s]"])
            return steps

        if process_name in ["Spin Coating", "Dip Coating", "Slot Die Coating", "Inkjet Printing"]: #Added DB2025-01-04 Printing
            steps = ["Material name", "Layer type", "Tool/GB name",]
            # Add solvent steps
            for i in range(1, config.get('solvents', 0) + 1):
                steps.extend([f"Solvent {i} name", f"Solvent {i} volume [uL]"])
            # Add solute steps
            for i in range(1, config.get('solutes', 0) + 1):
                steps.extend([f"Solute {i} type", f"Solute {i} Concentration [mM]"])
            # Add remaining steps based on the process type
            if process_name == "Spin Coating":
                steps.extend(["Solution volume [uL]", "Spin Delay [s]"])
                if config.get('spinsteps', 0) == 1:  # Added DB 2024-11-28 spinsteps
                    steps.extend(
                        ["Rotation speed [rpm]", "Rotation time [s]", "Acceleration [rpm/s]"])
                else:
                    for i in range(1, config.get('spinsteps', 0) + 1):  # Added DB 2024-11-28 spinsteps
                        steps.extend(
                            [f"Rotation speed {i} [rpm]", f"Rotation time {i} [s]", f"Acceleration {i} [rpm/s]"])

                if config.get("antisolvent", False):
                    steps.extend(["Anti solvent name", "Anti solvent volume [ml]", "Anti solvent dropping time [s]",
                                  "Anti solvent dropping speed [uL/s]", "Anti solvent dropping heigt [mm]"])  # Added DB 2024-11-28 speed and height

                if config.get("gasquenching", False):
                    steps.extend(["Gas", "Gas quenching start time [s]", "Gas quenching duration [s]", "Gas quenching flow rate [ml/s]", "Gas quenching pressure [bar]",
                                  "Gas quenching velocity [m/s]", "Gas quenching height [mm]", "Nozzle shape", "Nozzle size [mm²]"])  # Added DB 2024-11-29

                if config.get("vacuumquenching", False): # Added DB 2025-01-04
                    steps.extend(["Vacuum quenching start time [s]", "Vacuum quenching duration [s]", "Vacuum quenching pressure [bar]"
                                    ])  

            elif process_name == "Slot Die Coating": #Added DB2025-01-10/14 Added parameters/ coating run
                steps.extend(["Coating run", "Solution volume [um]", "Flow rate [uL/min]", "Head gap [mm]", "Speed [mm/s]",
                              "Air knife angle [°]", "Air knife gap [cm]", "Bead volume [mm/s]", "Drying speed [cm/min]",
                              "Drying gas temperature [°]", "Heat transfer coefficient [W m^-2 K^-1]", "Coated area [mm²]"])
                
            elif process_name == "Dip Coating":
                steps.extend(["Dipping duration [s]"]) #Added DB2025-01-10 Dip Coating
            
            elif process_name == "Inkjet Printing": #Added DB2025-01-14 Printing run
                steps.extend(["Printhead name", "Printing run", "Number of active nozzles", "Droplet density [dpi]", "Quality factor", "Step size", 
                              "Printing direction", "Printed area [mm²]", "Droplet per second [1/s]", "Droplet volume [pL]", "Dropping Height [mm]",
                              "wave function parameters 1", "wave function parameters 2", "wave function parameters 3", "wave function parameters 4",
                              "Ink reservoir pressure [bar]", "Table temperature [°C]", "Nozzle temperature [°C]", "Room temperature [°C]", "rel. humidity [%]" 
                              ]) # Added DB2025-01-08 more parameters
            
            # Add annealing steps
            steps.extend(["Annealing time [min]", "Annealing temperature [°C]",
                         "Annealing athmosphere", "Notes"])  # Added DB 2024-11-28 Atm
            return steps

        # PVD Processes multiple materials
        if process_name == "Seq-Evaporation" or process_name == "Co-Evaporation":
            steps = ["Material name", "Layer type", "Tool/GB name"]
            for i in range(1, config.get('materials', 0) + 1):
                steps.extend([f"Material name {i}", f"Base pressure {i} [bar]", f"Pressure start {i} [bar]", f"Pressure end {i} [bar]",
                              f"Source temperature start {i}[°C]", f"Source temperature end {i}[°C]", f"Substrate temperature {i} [°C]", 
                              f"Thickness {i} [nm]", f"Rate {i} [angstrom/s]", f"Tooling factor {i}"])
            return steps

        else:
            # Return predefined steps for other processes
            return config['steps']

    def adjust_column_widths(self):
        for col in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)  # Get the column letter
            for cell in col:
                if cell.value and isinstance(cell.value, str):
                    max_length = max(max_length, len(cell.value))
            adjusted_width = (max_length + 2)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def apply_nomad_id_formula(self):
        # Start from row 3 onward (assuming the first two rows are headers)
        for row in range(3, 30):  # Adjusted range for testing
            # The formula will be set for column F using Excel formula syntax
            # Changed DB 2024-11-28 DOES not Work...
            nomad_id_formula = f'=VERKETTEN("KIT_", B{row}, "_", A{row}, "_", C{row} ,"_", D{row}, "_", E{row})'
            self.worksheet[f"F{row}"].value = nomad_id_formula

    def save(self):
        current_date = datetime.now().strftime("%Y%m%d")
        filename = f"{current_date}_experiment_file.xlsx"
        self.workbook.save(filename)
        print(f"File saved as: {filename}")


# Define the base process configuration
process_config = {
    "Experiment Info": {"steps": ["Date", "Project_Name", "Batch", "Subbatch", "Sample", "Nomad ID", "Variation",
                                  "Sample dimension", "Sample area [cm^2]", "Number of pixels", "Pixel area", "Number of junctions", "Substrate material",
                                  "Substrate conductive layer", "Bottom Cell Name", "Notes"]}, # Added DB2025-01-08 Notes
    "Multijunction Info": {"steps": ["Recombination Layer", "Notes"]}, #Added DB2025-01-06 Multijunction Info, Added DB2025-01-08 Notes
    "Cleaning O2-Plasma": {"solvents": 1},
    "Cleaning UV-Ozone": {"solvents": 1},
    
    "Dip Coating": {"solvents": 1, "solutes": 1}, #Added DB2025-01-10 Adapted
    "Spin Coating": {"solvents": 1, "solutes": 1, "spinsteps": 1}, # Default values # Added DB 2024-11-28 spinsteps
    "Slot Die Coating": {"solvents": 1, "solutes": 1},  # Default values
    "Inkjet Printing": {"solvents": 1, "solutes": 1},  #Added DB2025-01-04 Printing

    "Evaporation": {"steps": ["Material name", "Layer type", "Tool/GB name", "Organic", "Base pressure [bar]", "Pressure start [bar]", "Pressure end [bar]",
                              "Source temperature start[°C]", "Source temperature end[°C]", "Substrate temperature [°C]", "Thickness [nm]",
                              "Rate [angstrom/s]", "Tooling factor", "Notes"]},  # Added DB 2024-11-29 many parameters #Could also be called Sublimation instead of Evaporation
    # Added DB 2024-11-29  multiple materials #Could also be called Co-Sublimation instead of Co-Evaporation
    "Co-Evaporation": {"materials": 2},
    # Added DB 2024-11-29  multiple materials #Could also be called Seq-Sublimation instead of Seq-Evaporation
    "Seq-Evaporation": {"materials": 2},
    "Close Space Sublimation": {"steps": ["Material name", "Layer type", "Tool/GB name", "Organic", "Process pressure [bar]",
                              "Source temperature [°C]", "Substrate temperature [°C]", "Material state", "Substrate source distance [mm]",
                              "Thickness [nm]", "Deposition Time [s]", "Carrier gas", "Notes"]},

    "Sputtering": {"steps": ["Material name", "Layer type", "Tool/GB name", "Gas", "Temperature [°C]", "Pressure [mbar]",
                             "Deposition time [s]", "Burn in time [s]", "Power [W]", "Rotation rate [rpm]",
                             "Thickness [nm]", "Gas flow rate [cm^3/min]", "Notes"]},
    "Laser Scribing": {"steps": ["Laser wavelength [nm]", "Laser pulse time [ps]", "Laser pulse frequency [kHz]",
                       "Speed [mm/s]", "Fluence [J/cm2]", "Power [%]", "Recipe file"]},
    "ALD": {"steps": ["Material name", "Layer type", "Tool/GB name", "Source", "Thickness [nm]", "Temperature [°C]", "Rate [A/s]",
                      "Time [s]", "Number of cycles", "Precursor 1", "Pulse duration 1 [s]",
                      "Manifold temperature 1 [°C]", "Bottle temperature 1 [°C]", "Precursor 2 (Oxidizer/Reducer)", "Pulse duration 2 [s]",
                      "Manifold temperature 2 [°C]"]},
    "Annealing": {"steps": ["Annealing time [min]", "Annealing temperature [°C]", "Annealing athmosphere", "Relative humidity [%]", "Notes"]}, # Added DB 2024-11-29 Annealing

    "Generic Process": {"steps": ["Name", "Notes"]}
}

# Define the sequence of processes with custom configurations

# Process for SJ antisolvent Daniel/Spinbot
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1,
#                                            "solutes": 1, "spinsteps": 1, "antisolvent": False}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2,
#                                            "solutes": 5, "spinsteps": 2, "antisolvent": True}},  # PSK
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2,
#                                            "spinsteps": 1, "antisolvent": False}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "ALD"},  # SnO2
#    {"process": "Evaporation"}  # Ag
#]


## Process for SJ Spinbot gasquenched Maria
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 5, "spinsteps": 2, "gasquenching": True}},  # PSK
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "Evaporation"}   # Ag
#]


# Process for SJ Vacuum Quenching Hang/Ting
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1,
#                                            "solutes": 1, "spinsteps": 1, "antisolvent": False}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2,
#                                            "solutes": 7, "spinsteps": 2, "vacuumquenching": True}},  # PSK vaccumquenching
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2,
#                                            "spinsteps": 1, "antisolvent": False}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "ALD"},  # SnO2
#    {"process": "Evaporation"}  # Ag
#]

## Process for SJ SD Kristina
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Slot Die Coating", "config":  {"solvents": 1, "solutes": 1}},  # SAM
#    {"process": "Evaporation"},                                                 # SAM Evap
#    {"process": "Slot Die Coating", "config":  {"solvents": 2, "solutes": 2}},  # PSK inorganic
#    {"process": "Slot Die Coating", "config":  {"solvents": 2, "solutes": 2}},  # PSK organic  
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "Evaporation"}  # Ag
#]

## Process for SJ 2step Ronja
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 3}},
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 4, "spinsteps": 2}},  # PSK inorganic
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 5, "spinsteps": 1}},  # PSK organic 
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 0, "spinsteps": 1}},  # PSK Washing
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "Evaporation"}   # Ag
#]


## Process for SJ Inkjet Raphael Theresa
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#    {"process": "Inkjet Printing", "config":  {"solvents": 2, "solutes": 2}},               # PSK inorganic
#    {"process": "Inkjet Printing", "config":  {"solvents": 1, "solutes": 3}},               # PSK organic 
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "Evaporation"}   # Ag
#]

## Process for SJ HybridInkjet Raphael 
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM Washing
#    {"process": "Seq-Evaporation", "config":  {"materials": 2} },                          # PSK inorganic
#    {"process": "Inkjet Printing", "config":  {"solvents": 1, "solutes": 3}},               # PSK organic 
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#   {"process": "ALD"},            #SnO2
#    {"process": "Evaporation"}   # Ag
#]

## Process for Tandem HybridInkjet Raphael 
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Multijunction Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Sputtering"},  #NiO
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1}},  # SAM
#    {"process": "Seq-Evaporation", "config":  {"materials": 2} },                          # PSK inorganic
#    {"process": "Inkjet Printing", "config":  {"solvents": 1, "solutes": 3}},               # PSK organic 
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "ALD"},            #SnO2
#    {"process": "Sputtering"},      #IZO
#    {"process": "Evaporation"},  # Ag
#    {"process": "Evaporation"}  # ARC
#]


## Process for Hybrid Ronja/Julian
#process_sequence = [
#   {"process": "Experiment Info"},
#   {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},   #SAM
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 0, "spinsteps":1 , "antisolvent": False}},   #SAM Washing
#   {"process": "Seq-Evaporation", "config":  {"materials": 2} },                                                  #PSK inorganic
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 4, "spinsteps":1 , "antisolvent": False}},   #PSK organic
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps":1 , "antisolvent": False}},   #Passivation Sol
#   {"process": "Evaporation"},    #Passivation Evap
#   {"process": "Evaporation"},    #C60
#   {"process": "Evaporation"},    #BCP
#   {"process": "ALD"},            #SnO2
#   {"process": "Evaporation"}     #Ag
# ]


## Process for Joshua Seq Evap
#process_sequence = [
#   {"process": "Experiment Info"},
#   {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},   #SAM
#   {"process": "Spin Coating", "config":  {"solvents": 1, "spinsteps":1 , "antisolvent": False}},   #SAM Washing
#   {"process": "Seq-Evaporation", "config":  {"materials": 3} },              #PSK inorganic
#   {"process": "Evaporation"},                                                #PSK organic
#   {"process": "Annealing"},      #Annealing
#   {"process": "Evaporation"},    #C60
#   {"process": "Evaporation"},    #BCP
#   {"process": "ALD"},            #SnO2
#   {"process": "Evaporation"}     #Ag
#]


## Process for  Co Evap
#process_sequence = [
#   {"process": "Experiment Info"},
#   {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},   #SAM
#   {"process": "Spin Coating", "config":  {"solvents": 1, "spinsteps":1 , "antisolvent": False}},   #SAM Washing
#   {"process": "Co-Evaporation", "config":  {"materials": 3} },              #PSK 
#   {"process": "Annealing"},      #Annealing
#   {"process": "Evaporation"},    #C60
#   {"process": "Evaporation"},    #BCP
#   {"process": "ALD"},            #SnO2
#   {"process": "Evaporation"}     #Ag
#]



## Process for CSS
#process_sequence = [
#   {"process": "Experiment Info"},
#   {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#   {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},   #SAM
#   {"process": "Spin Coating", "config":  {"solvents": 1, "spinsteps":1 , "antisolvent": False}},   #SAM Washing
#   {"process": "Close Space Sublimation"},              #PSK 
#   {"process": "Annealing"},      #Annealing
#   {"process": "Evaporation"},    #C60
#   {"process": "Evaporation"},    #BCP
#   {"process": "ALD"},            #SnO2
#   {"process": "Evaporation"}     #Ag
#]


## Process for Tandem SC Xuzheng Lingyi
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Multijunction Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 2}},
#    {"process": "Sputtering"},  
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}},  # NiO np
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps": 1, "antisolvent": False}},  # SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 7, "spinsteps": 2, "antisolvent": True}},  # PSK
#    {"process": "ALD"},           # AlO Passivation
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 2, "spinsteps": 1, "antisolvent": False}},  # Passivation Sol
#    {"process": "Evaporation"},  # Passivation Evap
#    {"process": "Evaporation"},  # C60
#    {"process": "Evaporation"},  # BCP
#    {"process": "ALD"},           # SnO2
#    {"process": "Sputtering"},      #IZO
#    {"process": "Evaporation"},  # Ag
#    {"process": "Evaporation"}  # ARC
#]



## Process for SOP
#process_sequence = [
#    {"process": "Experiment Info"},
#    {"process": "Cleaning O2-Plasma", "config": {"solvents": 3}},
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #NiO
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #SAM
#    {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 6, "spinsteps":2 , "antisolvent": True}}, #PSK
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #PEAI
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #PCBM
#    {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "antisolvent": False}}, #BCP
#    {"process": "Evaporation"} #Ag
# ]



# Test Process
# process_sequence = [
#     {"process": "Experiment Info"},
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1}},   #SAM
#     {"process": "Spin Coating", "config":  {"solvents": 2, "solutes": 5, "spinsteps":2 , "antisolvent": True}},    #PSK
#     {"process": "Spin Coating", "config":  {"solvents": 1, "solutes": 1, "spinsteps":1 , "gasquenching": True}},   #Passivation Sol
# ]


# Create an instance of ExperimentExcelBuilder and build the Excel file
process_sequence = []
builder = ExperimentExcelBuilder(process_sequence, process_config)
#builder.build_excel()
#builder.save()
