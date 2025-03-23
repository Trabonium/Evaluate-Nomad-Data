
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
import pandas as pd

from functions.api_calls_get_data import get_entryid, get_quantity_over_jv


### Function to get data from excel and server  ###_____________________________________________________________________________________

def get_data_excel_to_df(excel_file_path, nomad_url, token, key=["peroTF_CR_SpinBox_SpinCoating"], 
    columns_from_excel=[['sample_id', 5], ['variation', 6]]) -> pd.DataFrame:
    """columns_from excel: list of pairs of column name and column number (starting at 0) that will be read from the excel file.
    """
    
    #file_path = path + "ExperimentsInfo.xlsx"  
    
    workbook = load_workbook(filename=excel_file_path, data_only=True)
    sheet = workbook.active  # Access the active worksheet

    # Extract data from  specified columns, by default 6 (sample_id) and 7 (variation)
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        current_row = []
        for pair in columns_from_excel:
            current_row.append(row[pair[1]])
        data.append(current_row)

    # Convert the extracted data to a DataFrame
    column_titles = [pair[0] for pair in columns_from_excel]
    excel_df = pd.DataFrame(data, columns=column_titles)
    excel_df = excel_df.dropna(subset=["sample_id"])
    excel_df = excel_df[~excel_df["sample_id"].isin(["#NAME?", "KIT_____"])]
    
    df, quantities = get_batch_data(excel_df["sample_id"].unique().tolist(), nomad_url, token, key=key)
    # Merge with the existing DataFrame on 'sample_id'
    # Assume `df` is your existing DataFrame
    df = excel_df.merge(df, on="sample_id", how="left")

    # Print the updated DataFrame
    #print(df)
    return df


### Function to get data from the server and process it ###_________________________________________________________________________

def get_batch_data(sample_ids, nomad_url, token, quantities=["name"], key=["peroTF_CR_SpinBox_SpinCoating"]):
    #Get the NOMAD ID
    samples_of_batch = get_entryid(sample_ids, nomad_url, token)
    #samples_of_batch = [(sample_id, get_entryid(sample_id, nomad_url, token)) for sample_id in sample_ids]

    #Standard JV parameters to get
    jv_quantities=["efficiency",\
                   "fill_factor",\
                   "open_circuit_voltage",\
                   "short_circuit_current_density"]
    
    #Get data
    df = get_quantity_over_jv(samples_of_batch, key, quantities, jv_quantities, nomad_url, token)
    
    #Extract Information from ID
    #df['last_digit'] = df['sample_id'].str.extract('(\d)$').astype(int)[0]
    #df['category'] = df['sample_id'].str.split('_').str[4]
    #df['batch_name'] = df['sample_id'].str.split('_').str[3]
    #df['batch_date'] = df['sample_id'].str.split('_').str[2]

    # Data cleaning remove shunted or somehow damaged devices using filters
    #df = df[df['fill_factor'] >= 0.5]
    #df = df[df['open_circuit_voltage'] >= 1]
    #df = df[df['short_circuit_current_density'] <= 30]
    #df = df[df["fill_factor"] <= 1]


    return df, quantities






